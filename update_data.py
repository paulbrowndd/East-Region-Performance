#!/usr/bin/env python3
"""Update dashboard data.js from a daily Excel export.

Usage example:
  python3 update_data.py --xlsx "/Users/paul.brown/Downloads/4_24.xlsx" --date "4/24"
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install with: python3 -m pip install openpyxl"
    ) from exc


STATUS_KEYS = [
    "returns",
    "delayedArrival",
    "lateSort",
    "noSpokeScan",
    "incorrectFacility",
    "parcelPlannerMiss",
    "deliveredLate",
    "sortAppRollover",
    "lostAtHubTransport",
    "lostAtSpoke",
    "otdBy8pm",
    "autoReschedule",
    "siteRollover",
    "others",
]


def parse_date_key(value: str) -> Tuple[int, int]:
    match = re.fullmatch(r"\s*(\d{1,2})/(\d{1,2})\s*", value)
    if not match:
        raise ValueError(f"Invalid date key '{value}'. Expected M/D (example: 4/24).")
    return int(match.group(1)), int(match.group(2))


def safe_float(value) -> float:
    """Excel exports sometimes use string 'null' or blanks for empty numeric cells."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().lower()
    if s in ("", "null", "none", "n/a", "-", "na"):
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def load_data_js(path: Path) -> Dict:
    raw = path.read_text(encoding="utf-8").strip()
    prefix = "window.DASHBOARD_DATA="
    if not raw.startswith(prefix):
        raise ValueError("data.js does not start with window.DASHBOARD_DATA=")
    if raw.endswith(";"):
        raw = raw[:-1]
    return json.loads(raw[len(prefix) :])


def write_data_js(path: Path, payload: Dict) -> None:
    blob = "window.DASHBOARD_DATA=" + json.dumps(payload, separators=(",", ":")) + ";\n"
    path.write_text(blob, encoding="utf-8")


def status_breakdown(raw_status_counts: Dict[str, float]) -> Dict[str, int]:
    return {
        "returns": int(round(raw_status_counts.get("dispatched but return: return to dm", 0))),
        "delayedArrival": int(
            round(raw_status_counts.get("hub & spoke delay: delayed arrival at spoke", 0))
        ),
        "lateSort": int(round(raw_status_counts.get("late sort", 0))),
        "noSpokeScan": int(round(raw_status_counts.get("hub & spoke delay: no spoke scan", 0))),
        "incorrectFacility": int(
            round(raw_status_counts.get("hub & spoke delay: incorrect facility", 0))
        ),
        "parcelPlannerMiss": int(
            round(
                raw_status_counts.get(
                    "not included in parcel planner prior to first scan at spoke", 0
                )
            )
        ),
        "deliveredLate": int(round(raw_status_counts.get("delivered late: 12am - 4am", 0))),
        "sortAppRollover": int(
            round(sum(v for k, v in raw_status_counts.items() if "sort app rollover" in k))
        ),
        "lostAtHubTransport": int(
            round(raw_status_counts.get("hub & spoke delay: lost at hub or transportation", 0))
        ),
        "lostAtSpoke": int(round(raw_status_counts.get("hub & spoke delay: lost at spoke", 0))),
        "otdBy8pm": int(round(raw_status_counts.get("otd by 8pm", 0))),
        "autoReschedule": int(round(raw_status_counts.get("auto reschedule", 0))),
        "siteRollover": int(round(raw_status_counts.get("site rollover", 0))),
        "others": int(round(raw_status_counts.get("others", 0))),
    }


def parse_daily_xlsx(xlsx_path: Path, date_key: str) -> Dict:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)

    spokes_sheet = workbook["Spokes"]
    by_site: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    otd_by_site: Dict[str, float] = {}

    for row in spokes_sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        site = row[0]
        otd = row[2] if len(row) > 2 else None
        status = row[-2] if len(row) >= 2 else None
        mx_count = row[-1] if len(row) >= 1 else None
        if not site or not status:
            continue

        site_key = str(site).strip()
        status_key = str(status).strip().lower()
        by_site[site_key][status_key] += safe_float(mx_count)
        otd_by_site[site_key] = round(safe_float(otd) * 100, 1)

    spokes: List[Dict] = []
    for site in sorted(by_site.keys()):
        breakdown = status_breakdown(by_site[site])
        spokes.append({"code": site, "otd": otd_by_site[site], **breakdown})

    total_barcodes: Dict[str, float] = defaultdict(float)
    for status_counts in by_site.values():
        for status, count in status_counts.items():
            total_barcodes[status] += count
    barcode = [
        {"status": status, "count": int(round(count))}
        for status, count in sorted(total_barcodes.items(), key=lambda pair: (-pair[1], pair[0]))
    ]

    hubs_by_code: Dict[str, Dict] = {}
    current_hub = None
    missort_sheet = workbook["Hub Missorts"]
    for row in missort_sheet.iter_rows(min_row=2, values_only=True):
        col1, col2, col3 = row
        if isinstance(col1, str) and col1.strip():
            name = col1.strip()
            if name.lower().startswith("grand total"):
                current_hub = None
                continue
            current_hub = name
            hubs_by_code.setdefault(current_hub, {})
        if current_hub and isinstance(col2, str):
            metric = col2.strip().lower()
            if metric == "num missorts":
                hubs_by_code[current_hub]["missorts"] = int(round(safe_float(col3)))
            elif metric == "missort rate":
                hubs_by_code[current_hub]["missortRate"] = round(safe_float(col3) * 100, 3)

    cpt_sheet = workbook["Hub On Time CPT"]
    for row in cpt_sheet.iter_rows(min_row=3, values_only=True):
        hub_code, value = row
        if not hub_code:
            continue
        key = str(hub_code).strip()
        if key.lower().startswith("grand total"):
            continue
        hubs_by_code.setdefault(key, {})
        hubs_by_code[key]["onTimeCpt"] = round(safe_float(value) * 100, 1)

    hubs = [
        {
            "code": code,
            "missorts": values.get("missorts", 0),
            "missortRate": values.get("missortRate", 0),
            "onTimeCpt": values.get("onTimeCpt", 0),
        }
        for code, values in sorted(hubs_by_code.items())
    ]

    # Parse Lanes That Missed CPT sheet if present
    cpt_lanes = []
    if "Lanes That Missed CPT" in workbook.sheetnames:
        from datetime import timedelta

        def fmt_time_est(dt) -> str:
            if not hasattr(dt, 'hour'):
                return "N/A"
            est = dt - timedelta(hours=4)
            hour = est.hour % 12 or 12
            ampm = "AM" if est.hour < 12 else "PM"
            return f"{hour}:{est.minute:02d} {ampm}"

        def fmt_delta(actual, scheduled) -> str:
            if not hasattr(actual, 'hour') or not hasattr(scheduled, 'hour'):
                return ""
            diff = int((actual - scheduled).total_seconds() / 60)
            if diff > 0:
                return f"+{diff}m late"
            elif diff < 0:
                return f"{abs(diff)}m early"
            return "On time"

        lanes_sheet = workbook["Lanes That Missed CPT"]
        for row in lanes_sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            hub, lane, shipment_id, pro, sched_pickup, actual_pickup, cpt_time, departure = row[:8]
            cpt_lanes.append({
                "hub": str(hub).strip(),
                "lane": str(lane).strip() if lane else "",
                "shipmentId": str(shipment_id).strip() if shipment_id else "",
                "scheduledPickup": fmt_time_est(sched_pickup),
                "actualPickup": fmt_time_est(actual_pickup),
                "pickupStatus": fmt_delta(actual_pickup, sched_pickup),
                "cpt": fmt_time_est(cpt_time),
                "departure": fmt_time_est(departure),
                "departureStatus": fmt_delta(departure, cpt_time),
            })

    return {"label": f"Daily {date_key}", "spokes": spokes, "hubs": hubs, "barcode": barcode, "cptLanes": cpt_lanes}


def build_weekly_snapshot(daily_periods: List[Dict], week_label: str) -> Dict:
    spokes_by_code: Dict[str, List[Dict]] = defaultdict(list)
    for period in daily_periods:
        for spoke in period.get("spokes", []):
            spokes_by_code[spoke["code"]].append(spoke)

    weekly_spokes: List[Dict] = []
    for code, rows in sorted(spokes_by_code.items()):
        item = {"code": code, "otd": round(sum(r.get("otd", 0) for r in rows) / len(rows), 1)}
        for key in STATUS_KEYS:
            item[key] = sum(r.get(key, 0) for r in rows)
        weekly_spokes.append(item)

    hubs_by_code: Dict[str, List[Dict]] = defaultdict(list)
    for period in daily_periods:
        for hub in period.get("hubs", []):
            hubs_by_code[hub["code"]].append(hub)

    weekly_hubs: List[Dict] = []
    for code, rows in sorted(hubs_by_code.items()):
        weekly_hubs.append(
            {
                "code": code,
                "missorts": sum(r.get("missorts", 0) for r in rows),
                "missortRate": round(sum(r.get("missortRate", 0) for r in rows) / len(rows), 3),
                "onTimeCpt": round(sum(r.get("onTimeCpt", 0) for r in rows) / len(rows), 1),
            }
        )

    barcode_totals: Dict[str, int] = defaultdict(int)
    for period in daily_periods:
        for row in period.get("barcode", []):
            barcode_totals[row["status"]] += int(row["count"])
    weekly_barcode = [
        {"status": status, "count": count}
        for status, count in sorted(barcode_totals.items(), key=lambda pair: (-pair[1], pair[0]))
    ]

    return {
        "label": week_label,
        "spokes": weekly_spokes,
        "hubs": weekly_hubs,
        "barcode": weekly_barcode,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Update data.js with a new daily Excel file.")
    parser.add_argument("--xlsx", required=True, help="Path to daily Excel file.")
    parser.add_argument("--date", required=True, help="Date key in M/D format (example: 4/24).")
    parser.add_argument(
        "--data-file",
        default="data.js",
        help="Path to data.js (default: ./data.js).",
    )
    parser.add_argument(
        "--week-key",
        default=None,
        help='Weekly snapshot key to update (default: first existing key or "Week of <date>").',
    )
    parser.add_argument(
        "--week-dates",
        default=None,
        help="Comma-separated daily date keys to include in weekly rollup (default: all daily snapshots).",
    )
    parser.add_argument(
        "--week-label",
        default="Weekly (Mon-Sun)",
        help='Weekly label text (default: "Weekly (Mon-Sun)").',
    )
    args = parser.parse_args()

    parse_date_key(args.date)

    data_path = Path(args.data_file).expanduser().resolve()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    payload = load_data_js(data_path)

    payload.setdefault("dailySnapshots", {})
    payload.setdefault("weeklySnapshots", {})

    daily_snapshot = parse_daily_xlsx(xlsx_path, args.date)
    payload["dailySnapshots"][args.date] = daily_snapshot

    sorted_daily_keys = sorted(payload["dailySnapshots"].keys(), key=parse_date_key)
    payload["dailySnapshots"] = {key: payload["dailySnapshots"][key] for key in sorted_daily_keys}

    if args.week_dates:
        week_dates = [piece.strip() for piece in args.week_dates.split(",") if piece.strip()]
        for key in week_dates:
            parse_date_key(key)
    else:
        week_dates = sorted_daily_keys

    missing = [key for key in week_dates if key not in payload["dailySnapshots"]]
    if missing:
        raise ValueError(f"Weekly rollup dates not found in daily snapshots: {missing}")

    weekly_key = args.week_key
    if not weekly_key:
        existing = list(payload["weeklySnapshots"].keys())
        weekly_key = existing[0] if existing else f"Week of {args.date}"

    selected_daily = [payload["dailySnapshots"][key] for key in week_dates]
    payload["weeklySnapshots"][weekly_key] = build_weekly_snapshot(selected_daily, args.week_label)

    write_data_js(data_path, payload)
    print(f"Updated {data_path}")
    print(f"- Daily snapshot: {args.date} ({len(daily_snapshot['spokes'])} spokes)")
    print(f"- Weekly snapshot: {weekly_key} from {', '.join(week_dates)}")


if __name__ == "__main__":
    main()
